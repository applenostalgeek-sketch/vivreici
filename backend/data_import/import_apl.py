"""
Import de l'APL (Accessibilité Potentielle Localisée) aux médecins généralistes.

Source : DREES — data.drees.solidarites-sante.gouv.fr
Millésime : 2023 (le plus récent disponible)

L'APL mesure le nombre de consultations/visites accessibles par habitant standardisé,
en tenant compte de l'offre et de la demande dans un rayon de chalandise (pas seulement
les médecins de la commune). Indicateur bien plus pertinent que nb_medecins/10000.

Lancement :
  cd /Users/admin/vivreici && .venv/bin/python -m backend.data_import.import_apl
"""

import asyncio
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

import httpx
import openpyxl
import pandas as pd

from sqlalchemy import text
from backend.database import async_session, init_db
from backend.scoring import calculer_score_global, percentile_to_score


APL_URL = (
    "https://data.drees.solidarites-sante.gouv.fr/api/v2/catalog/datasets"
    "/530_l-accessibilite-potentielle-localisee-apl/attachments"
    "/indicateur_d_accessibilite_potentielle_localisee_apl_aux_medecins_generalistes_xlsx"
)


async def telecharger_apl() -> dict[str, float]:
    """Télécharge le fichier DREES APL et retourne {code_insee: apl_value}."""
    print("Téléchargement APL médecins généralistes (DREES)...")
    async with httpx.AsyncClient(follow_redirects=True, timeout=180) as client:
        resp = await client.get(APL_URL)
        resp.raise_for_status()
    print(f"  → {len(resp.content) / 1024:.0f} Ko reçus")

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)

    # Utiliser la feuille la plus récente (APL 2023 > APL 2022)
    sheet_candidates = [s for s in wb.sheetnames if s.startswith("APL")]
    sheet_name = max(sheet_candidates) if sheet_candidates else wb.sheetnames[0]
    print(f"  → Feuille utilisée : {sheet_name}")

    ws = wb[sheet_name]
    data: dict[str, float] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 11:  # header aux lignes 9-10, données à partir de la ligne 11
            continue
        code, _, apl = row[0], row[1], row[2]
        if code and isinstance(code, str) and len(code) == 5 and apl is not None:
            try:
                data[code] = float(apl)
            except (ValueError, TypeError):
                continue
    wb.close()
    print(f"  → {len(data):,} communes avec APL")
    return data


async def run():
    print("=== Import APL (Accessibilité Potentielle Localisée) ===\n")
    await init_db()

    # 1. S'assurer que la colonne apl_medecins existe (migration légère)
    async with async_session() as session:
        try:
            await session.execute(text("ALTER TABLE scores ADD COLUMN apl_medecins REAL DEFAULT -1"))
            await session.commit()
            print("  → Colonne apl_medecins ajoutée")
        except Exception:
            pass  # Colonne déjà présente

    # 2. Télécharger les données APL
    apl_data = await telecharger_apl()

    if not apl_data:
        print("ERREUR : aucune donnée APL récupérée.")
        return

    # 3. Calculer les scores par percentile national
    df = pd.DataFrame(apl_data.items(), columns=["code_insee", "apl"])
    serie = df["apl"]

    df["score_sante"] = df["apl"].apply(
        lambda x: round(percentile_to_score(x, serie, "direct"), 1)
    )

    print(f"\nStats APL :")
    print(f"  Médiane   : {serie.median():.2f} consultations/an/hab")
    print(f"  P10 / P90 : {serie.quantile(0.10):.2f} / {serie.quantile(0.90):.2f}")
    print(f"  Score santé médian : {df['score_sante'].median():.1f}")

    # 4. Mise à jour en base
    print("\nMise à jour score_sante + apl_medecins...")
    async with async_session() as session:
        nb_ok = 0
        for _, row in df.iterrows():
            res = await session.execute(text("""
                UPDATE scores
                SET score_sante  = :score,
                    apl_medecins = :apl,
                    updated_at   = CURRENT_TIMESTAMP
                WHERE code_insee = :code
            """), {"score": float(row["score_sante"]), "apl": float(row["apl"]), "code": row["code_insee"]})
            if res.rowcount > 0:
                nb_ok += 1
            if nb_ok % 5000 == 0:
                await session.commit()
                print(f"  → {nb_ok:,} communes mises à jour...")
        await session.commit()
    print(f"  → {nb_ok:,} communes mises à jour")

    # 5. Recalcul des scores globaux
    print("\nRecalcul des scores globaux...")
    async with async_session() as session:
        result = await session.execute(text("""
            SELECT code_insee,
                   score_equipements, score_securite, score_immobilier,
                   score_education,   score_sante,     score_revenus,
                   score_transports,  score_demographie, score_environnement
            FROM scores
            WHERE score_equipements >= 0 OR score_securite    >= 0
               OR score_immobilier  >= 0 OR score_education   >= 0
               OR score_sante       >= 0 OR score_transports  >= 0
               OR score_demographie >= 0 OR score_environnement >= 0
        """))
        rows = result.fetchall()
        cols = ["code_insee", "score_equipements", "score_securite", "score_immobilier",
                "score_education", "score_sante", "score_revenus",
                "score_transports", "score_demographie", "score_environnement"]
        cat_map = {
            "score_equipements":   "equipements",
            "score_securite":      "securite",
            "score_immobilier":    "immobilier",
            "score_education":     "education",
            "score_sante":         "sante",
            # score_revenus exclu intentionnellement (cohésion sociale supprimée du scoring)
            "score_transports":    "transports",
            "score_demographie":   "demographie",
            "score_environnement": "environnement",
        }
        nb_recalc = 0
        for row in rows:
            r = dict(zip(cols, row))
            sous_scores = {
                cat: r[col] for col, cat in cat_map.items()
                if r[col] is not None and r[col] >= 0
            }
            if not sous_scores:
                continue
            score, lettre, nb = calculer_score_global(sous_scores)
            await session.execute(text("""
                UPDATE scores
                SET score_global = :sg, lettre = :l,
                    nb_categories_scorees = :nb, updated_at = CURRENT_TIMESTAMP
                WHERE code_insee = :c
            """), {"sg": score, "l": lettre, "nb": nb, "c": r["code_insee"]})
            nb_recalc += 1
            if nb_recalc % 5000 == 0:
                await session.commit()
        await session.commit()
        print(f"  → {nb_recalc:,} scores globaux recalculés")

    print("\n=== Import APL terminé ===")


if __name__ == "__main__":
    asyncio.run(run())
